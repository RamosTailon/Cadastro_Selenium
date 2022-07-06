''' 
BOT PARA COLHER INFORMAÇÕES DO SISTEMA DE CADASTRO

'''

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from time import sleep
import re

navegador = webdriver.Chrome()
pagina_autenticacao = "https://valparaisodegoias.camaravirtual.app.br/autenticacao"
pagina_user = "https://valparaisodegoias.camaravirtual.app.br/admin/cadastro/usuario"

log = "login"
sen = "senha"

navegador.get(pagina_autenticacao)

sleep(5)

insert_login = str(input("Digite seu Login: "))
insert_senha = str(input("Digite sua Senha: "))


navegador.find_element(By.NAME,log).send_keys(insert_login)
navegador.find_element(By.NAME,sen).send_keys(insert_senha)
navegador.find_element(By.NAME,sen).send_keys(Keys.RETURN)

sleep(10)
navegador.get(pagina_user)

''' 
VERIFICADOR DE HORÁRIO
    Verifica o horário que começou a rodar o programa para apresentar Feedback
'''
from datetime import datetime
now = datetime.now()
hora = now.hour
minuto = now.minute

''' 
ADICIONA O SHEET DA PLANILHA
    Adiciona o openpyxl para adicionar as informações em tabela
'''

import openpyxl

book = openpyxl.Workbook()

book.create_sheet('Pessoas')

#função para tratar os dados em string e formata-los
#formata apenas conteúdo com números


import openpyxl

book = openpyxl.Workbook()

book.create_sheet('Pessoas')

def numeros_telefone(f):
    try:
        #RE.SEARCH -> pega informação do caractere : até o Caractere R e guarda na variável icod
        icod = re.search(r':.*R',f)
        p = icod.span()[0] +1
        q = icod.span()[1] -1
        return f[p:q]
    except:
        #RE.SEARCH -> pega informação do caractere : até o Caractere N e guarda na variável icod
        icod = re.search(r':.*N',f)
        p = icod.span()[0] +1
        q = icod.span()[1] -1
        return f[p:q]
        
def sem_numero(f):
    try:
        #se nao houver número ele retorna sem numero
        return 'sem número'
    except:
        return 'nao orientado'


lista_de_pessoas = []
#nome,cpf,email,data de nascimento,telefone
pessoas = []

#*******************************************************

#Inicio e Fim do Loop
inicio = int(input("Digite o primeiro ID: "))
quant_pessoas = int(input("Digite o último ID: "))


''' 
PEGAR AS INFORMAÇÕES
        Essa função verificador() pega as informações e guarda em um lista dentro de lista
'''
def verificador(id):
    sleep(1)

    navegador.find_element(By.XPATH, '//*[@id="containerfluid"]/div/div/div[2]/div/div/div[1]/div[2]/div/div/input').send_keys(f"{id}")
    
    sleep(3)
    
    #NOME PESSOA
    nome = navegador.find_element(By.XPATH, '//*[@id="nome"]').get_attribute("value")
    
    #CPF PESSOA
    cpf = navegador.find_element(By.XPATH, '//*[@id="cpf"]').get_attribute("value")
    
    #EMAIL PESSOA
    email = navegador.find_element(By.XPATH, '//*[@id="email"]').get_attribute("value")
    
    #DATA NASCIMENTO PESSOA
    data = navegador.find_element(By.XPATH, '//*[@id="documentacao_pessoal.data_nascimento"]').get_attribute("value")
    
    #TELEFONE PESSOA
    telefone_pessoa = navegador.find_element(By.XPATH, '//*[@id="containers"]/div[2]/div').text.replace('\n', '')

    veri = any(map(str.isdigit,telefone_pessoa))
    if veri == True:
        telefone = numeros_telefone(telefone_pessoa)
    else:
        telefone = sem_numero(telefone_pessoa)
    
    pessoas.append(nome)
    pessoas.append(cpf)
    pessoas.append(email)
    pessoas.append(data)
    pessoas.append(telefone)
    lista_de_pessoas.append(pessoas[:])
    print(f'cadastro {id} de {quant_pessoas -1}')
    pessoas.clear()
    
    navegador.find_element(By.XPATH, '//*[@id="containerfluid"]/div/div/div[2]/div/div/div[1]/div[2]/div/div/input').clear()
    

#*****************************************************

''' 
LOOP
    O Loop faz a interação for do primeiro id ao último
    FeedBack apresentados:
        >Inicio do processo
        >Quantidade de loop feito
        >Quantidade de cadastros realizados
'''

print(f'INICIO -> {hora}:{minuto}')
for pi in range(inicio,quant_pessoas):
    verificador(pi)

print(f'{quant_pessoas-1-inicio} CADASTROS REALIZADOS!')
#*******************************************************

''' 
PLANILHA COM OPENPYXL
    Adiciona as listas em uma variável para adicionar um sheet
'''
planilha_nomes = book['Pessoas']

for i in range(len(lista_de_pessoas)):
    planilha_nomes.append(lista_de_pessoas[i])
#********************************************************

''' 
EXCEL
    Salva em formato excel (xlsx)
'''
book.save('Camara_Virtual.xlsx')